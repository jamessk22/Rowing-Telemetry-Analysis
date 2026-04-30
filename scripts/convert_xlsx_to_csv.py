#!/usr/bin/env python3
from __future__ import annotations
from pathlib import Path
import argparse
import csv
import re
import sys


try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - handled for user guidance
    raise SystemExit(
        "Missing dependency: openpyxl.\n"
        "Install openpyxl in your active Python environment before running this script."
    ) from exc


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Recursively convert .xlsx workbooks to .csv files. By default, the script "
            "writes CSVs next to the source workbooks."
        )
    )
    parser.add_argument("input_dir", type=Path, help="Root folder to scan for .xlsx files.")
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        help=(
            "Destination root for generated CSV files. Defaults to the input directory, "
            "which writes CSVs next to the source workbooks."
        ),
    )
    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="Export every worksheet in each workbook. Default: only the first sheet.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing CSV files. Default: skip files that already exist.",
    )
    parser.add_argument(
        "--delete-xlsx",
        action="store_true",
        help=(
            "Delete each source .xlsx file after its CSV export succeeds. "
            "Use with care."
        ),
    )
    return parser.parse_args()


def iter_workbooks(root: Path) -> list[Path]:
    return sorted(path for path in root.rglob("*.xlsx") if path.is_file())


def build_output_path(
    workbook_path: Path,
    sheet_name: str,
    input_root: Path,
    output_root: Path,
    include_sheet_name: bool,
) -> Path:
    relative_path = workbook_path.relative_to(input_root)
    target = output_root / relative_path.with_suffix(".csv")

    if include_sheet_name:
        sanitized = re.sub(r"[^A-Za-z0-9._-]+", "-", sheet_name.strip()).strip("-") or "sheet"
        target = target.with_name(f"{target.stem}__{sanitized}{target.suffix}")

    return target


def write_sheet_to_csv(worksheet, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])


def convert_workbook(
    workbook_path: Path,
    input_root: Path,
    output_root: Path,
    export_all_sheets: bool,
    overwrite: bool,
    delete_xlsx: bool,
) -> tuple[int, int, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheets = workbook.worksheets if export_all_sheets else workbook.worksheets[:1]
    include_sheet_name = export_all_sheets and len(workbook.worksheets) > 1
    written = 0
    skipped = 0
    deleted = 0

    try:
        for worksheet in worksheets:
            output_path = build_output_path(
                workbook_path=workbook_path,
                sheet_name=worksheet.title,
                input_root=input_root,
                output_root=output_root,
                include_sheet_name=include_sheet_name,
            )

            if output_path.exists() and not overwrite:
                skipped += 1
                print(f"SKIP {output_path}")
                continue

            write_sheet_to_csv(worksheet, output_path)
            written += 1
            print(f"WROTE {output_path}")
    finally:
        workbook.close()

    if delete_xlsx and written > 0:
        workbook_path.unlink()
        deleted = 1
        print(f"DELETED {workbook_path}")

    return written, skipped, deleted


def main() -> int:
    args = parse_args()
    input_root = args.input_dir.expanduser().resolve()
    output_root = (args.output_dir or args.input_dir).expanduser().resolve()

    if not input_root.exists():
        print(f"Input folder does not exist: {input_root}", file=sys.stderr)
        return 1
    if not input_root.is_dir():
        print(f"Input path is not a folder: {input_root}", file=sys.stderr)
        return 1

    workbooks = iter_workbooks(input_root)
    if not workbooks:
        print(f"No .xlsx files found under {input_root}")
        return 0

    total_written = 0
    total_skipped = 0
    total_deleted = 0
    for workbook_path in workbooks:
        written, skipped, deleted = convert_workbook(
            workbook_path=workbook_path,
            input_root=input_root,
            output_root=output_root,
            export_all_sheets=args.all_sheets,
            overwrite=args.overwrite,
            delete_xlsx=args.delete_xlsx,
        )
        total_written += written
        total_skipped += skipped
        total_deleted += deleted

    print(
        f"Done. Converted {total_written} sheet(s) from {len(workbooks)} workbook(s); "
        f"skipped {total_skipped} existing file(s); deleted {total_deleted} source workbook(s)."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
